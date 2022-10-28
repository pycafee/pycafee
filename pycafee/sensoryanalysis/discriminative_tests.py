"""This module focuses on discriminative sensory analysis tests

These tests are usually separated into two groups (difference tests and sensitivity test), but else will be added as a single group in this module. Each class will have a single test.

Difference tests:

- paired comparison;
- triangular;
- duo-trio;
- multiple comparison;
- ordering;
- A or Not-A;
- two out of five.

Sensitivity tests:

- Limits;
- constant stimulation;
- dilution.


"""


#########################################
################ Imports ################
#########################################

###### Standard ######
from collections import namedtuple
from datetime import datetime
from math import ceil

###### Third part ######
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
import scipy.stats as stats
from scipy.stats import binom
from tabulate import tabulate



###### Home made ######


from pycafee.utils import helpers
from pycafee.utils import general
from pycafee.utils import checkers

from pycafee.utils.helpers import AlphaManagement, NDigitsManagement, LanguageManagement



from pycafee.database_management import management

###########################################
################ Functions ################
###########################################



class TriangleTest(AlphaManagement, NDigitsManagement):


    def __init__(self, alfa=None, language=None, n_digits=None, **kwargs):
        super().__init__(alfa=alfa, language=language, n_digits=n_digits, **kwargs)
        self.df_table = None
        self.conclusion = None




    # with tests, with text, with database, with docstring
    def get_number_assessors(self, pd=None, beta=None, alfa=None):
        """This function returns the minimum number of assessors to perform the triangular test [1]_.

        Parameters
        ----------
        pd : ``str``, optional
            The proportion of the population represented by the assessors that can distinguish between the two products. It can be:

            * ``50%``;
            * ``40%``;
            * ``30%`` (default);
            * ``20%``;
            * ``10%``;


        alfa : ``float``, optional
            The probability of concluding that a perceptible difference exists when, in reality, one does not (Type I Error). It can be:

            * ``0.20``;
            * ``0.10``;
            * ``0.05`` (default);
            * ``0.01``;
            * ``0.001``;


        beta : ``float``, optional
            The probability of concluding that no perceptible difference exists when, in reality, one does (Type II Error). It can be:

            * ``0.20``;
            * ``0.10``;
            * ``0.05`` (default);
            * ``0.01``;
            * ``0.001``;


        Returns
        -------
        n_assessors : ``int``
            The minimum number of assessors to perform the triangular test
        inputs : ``dict``
            A dictionary with the parameters used to obtain the minimum value of assessors


        References
        ----------
        .. [1] Standard Test Method for Sensory Analysis—Triangle Test, Designation: E1885 − 04, 2011.


        Examples
        --------

        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, input = tringular.get_number_assessors()
        >>> print(result)
        TriangleTest(minimum_number_of_assessors=66)
        >>> print(input)
        {'pd': '30%', 'alfa': 0.05, 'beta': 0.05}


        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, input = tringular.get_number_assessor(pd="10%", alfa=0.01)
        >>> print(result)
        TriangleTest(minimum_number_of_assessors=824)
        >>> print(input)
        {'pd': '10%', 'alfa': 0.01, 'beta': 0.05}


        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest(language="pt-br")
        >>> result, input = tringular.get_number_assessors(pd="20%", alfa=0.05, beta=0.10)
        >>> print(result)
        TesteTriangular(numero_minimo_de_avaliadores=117)
        >>> print(input)
        {'pd': '20%', 'alfa': 0.05, 'beta': 0.1}


        """

        # ----- checking alpha value ----- #
        if alfa is None:
            alfa = self.alfa
        else:
            # --- should be float --- #
            checkers._check_is_float(alfa, "alfa", self.language)
            # --- should be in the range (0,1) --- #
            checkers._check_data_in_range(alfa, "alfa", 0.0, 1.0, self.language)
            # --- must be in --- #
            alloweds = [0.20, 0.10, 0.05, 0.01, 0.001]
            helpers._raises_wrong_param("alfa", alloweds, alfa, self.language)

        # ----- checking beta value ----- #
        if beta is None:
            beta = 0.05
        else:
            # --- should be float --- #
            checkers._check_is_float(beta, "beta", self.language)
            # --- should be in the range (0,1) --- #
            checkers._check_data_in_range(beta, "beta", 0.0, 1.0, self.language)
            # --- must be in --- #
            alloweds = [0.20, 0.10, 0.05, 0.01, 0.001]
            helpers._raises_wrong_param("beta", alloweds, beta, self.language)


        # ----- checking the pd value ----- #
        if pd is None:
            pd = "30%"
        else:
            # --- should be str --- #
            checkers._check_is_str(pd, "pd", language=self.language)
            # --- must be in --- #
            alloweds = ["10%", "20%", "30%", "40%", "50%"]
            helpers._raises_wrong_param("pd", alloweds, pd, self.language)

        fk_id_function = management._query_func_id("discriminative_tests")
        messages = management._get_messages(fk_id_function, self.language, "discriminative_tests")
        # ----- grouping entries into a dictionary ----- #
        inputs = {
            "pd": pd,
            messages[1][2][0]: alfa,
            messages[1][3][0]: beta
            }

        # ----- getting the position for each beta value ----- #
        if beta == 0.20:
            position = 0
        elif beta == 0.10:
            position = 1
        elif beta == 0.05:
            position = 2
        elif beta == 0.01:
            position = 3
        else:
            position = 4

        # ----- dictionary with the data ----- #
        n_assessors = {
            "50%": {
                0.20: [7, 12, 16, 25, 36],
                0.10: [12, 15, 20, 30, 43],
                0.05: [16, 20, 23, 35, 48],
                0.01: [25, 30, 35, 47, 62],
                0.001: [36, 43, 48, 62, 81],
            },
            "40%": {
                0.20: [12, 17, 25, 36, 55],
                0.10: [17, 25, 30, 46, 67],
                0.05: [23, 30, 40, 57, 79],
                0.01: [35, 47, 56, 76, 102],
                0.001: [55, 68, 76, 102, 130]
            },
            "30%": {
                0.20: [20, 28, 39, 64, 97],
                0.10: [30, 43, 54, 81, 119],
                0.05: [40, 53, 66, 98, 136],
                0.01: [62, 82, 97, 131, 181],
                0.001: [93, 120, 138, 181, 233]
            },
            "20%": {
                0.20: [39, 64, 86, 140, 212],
                0.10: [62, 89, 119, 178, 260],
                0.05: [87, 117, 147, 213, 305],
                0.01: [136, 176, 211, 292, 397],
                0.001: [207, 257, 302, 396, 513]
            },
            "10%": {
                0.20: [149, 238, 325, 529, 819],
                0.10: [240, 348, 457, 683, 1011],
                0.05: [325, 447, 572, 828, 1181],
                0.01: [525, 680, 824, 1132, 1539],
                0.001: [803, 996, 1165, 1530, 1992]
            }
        }

        # ----- getting the results ----- #
        minimum_number_of_assessors = n_assessors[pd][alfa][position]

        result = namedtuple(messages[1][0][0], (messages[1][1][0],))
        return result(minimum_number_of_assessors), inputs


    # with tests, with text, with database, with docstring
    def minimum_of_correct_responses(self, n_of_assessors, alfa=None):
        """This functions returns the minimum number of correct responses required for significance at the stated α level for the corresponding number of assessors.

        Parameters
        ----------
        n_of_assessors : ``int``
            The number of assessors (greater than 6)

        alfa : ``float``, optional
            The significance level. It can be:

            * ``0.20``;
            * ``0.10``;
            * ``0.05`` (default);
            * ``0.01``;
            * ``0.001``;


        Returns
        -------
        min_responses : ``int``
            The minimum number of correct responses required for significance at the stated α level
        inputs : ``dict``
            A dictionary with the parameters used to obtain the result


        Notes
        -----
        The minimum number of correct answers is obtained from the Binomial distribution using function ``scipy.stats.binom.ppf()`` [1]_ as follows:

        .. code:: python

            int(binom.ppf(1-alfa, n_of_assessors, 1/3) + 1)



        References
        ----------
        .. [1] SCIPY. scipy.stats.binom. Available at: `www.scipy.org <https://docs.scipy.org/doc/scipy/reference/generated/scipy.stats.binom.html>`_. Access on: 22 Jul. 2022.


        Examples
        --------

        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, inputs = tringular.minimum_of_correct_responses(10)
        >>> print(result)
        TriangleTest(minimum_of_correct_responses=7)
        >>> print(inputs)
        {'alpha': 0.05, 'n_of_assessors': 10}


        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest(language="pt-br")
        >>> result, inputs = tringular.minimum_of_correct_responses(36)
        >>> print(result)
        TesteTriangular(minimo_de_respostas_corretas=18)
        >>> print(inputs)
        {'alfa': 0.05, 'n_of_assessors': 36}


        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, inputs = tringular.minimum_of_correct_responses(36, alfa = 0.01)
        >>> print(result)
        TriangleTest(minimum_of_correct_responses=20)
        >>> print(inputs)
        {'alpha': 0.01, 'n_of_assessors': 36}


        """
        # ----- checking n_of_assessors value ----- #
        checkers._check_is_integer(n_of_assessors, "n_of_assessors", self.language)
        checkers._check_value_is_equal_or_higher_than(n_of_assessors, "n_of_assessors", 6, language=self.language)

        # ----- checking alpha value ----- #
        if alfa is None:
            alfa = self.alfa
        else:
            # --- should be float --- #
            checkers._check_is_float(alfa, "alfa", self.language)
            # --- should be in the range (0,1) --- #
            checkers._check_data_in_range(alfa, "alfa", 0.0, 1.0, self.language)
            # --- must be in --- #
            alloweds = [0.20, 0.10, 0.05, 0.01, 0.001]
            helpers._raises_wrong_param("alfa", alloweds, alfa, self.language)


        fk_id_function = management._query_func_id("discriminative_tests")
        messages = management._get_messages(fk_id_function, self.language, "discriminative_tests")

        # ----- getting critical value from binomial distribution with p=1/3 ----- #
        min_responses = int(binom.ppf(1-alfa, n_of_assessors, 1/3) + 1)

        inputs = {
            messages[1][2][0]: alfa,
            "n_of_assessors": n_of_assessors,
        }

        result = namedtuple(messages[1][0][0], (messages[1][4][0],))
        return result(min_responses), inputs


    # with tests, with text, with database, with docstring
    def make_combinations(self, n_of_assessors, seed=None, shuffle=None, reorder=None):
        """This function creates a dataframe with all the combinations to apply the triangle test.


        Parameters
        ----------
        n_of_assessors : ``int``
            The number of assessors (greater than 6)
        seed : ``int`` (positive), optional
            The seed that controls the randomness of the output.

            * If ``None`` (default), three-digit numbers are randomly generated without a specific seed;
            * If a number, three-digit numbers are generated using randomness specified by the seed (which can be used to replicate the random results).


        shuffle : ``bool``, optional
            Whether permutations between ``"A"`` and ``"B"`` should be randomized or not.

            * If ``True`` (default), the six possible combinations will be randomized. The ``seed`` parameter controls the randomness of this parameter;
            * If ``False``, the combinations follow the following pattern: ``"AAB", "ABA", "ABB", "BBA", "BAB", "BAA"``.

        reorder : ``bool``, optional
            Whether the table should should be randomized or not.

            * If ``True``, the table will be randomized. The ``seed`` parameter controls the randomness of this parameter;
            * If ``False`` (default), the table will keep the original pattern.


        Returns
        -------
        min_responses : ``pd.DataFrame``
            The dataframe with all the combinations required to prepare the Triangle Test
        inputs : ``dict``
            A dictionary with the parameters used to obtain the table



        References
        ----------
        .. [1] Standard Test Method for Sensory Analysis—Triangle Test, Designation: E1885 − 04, 2011.


        Examples
        --------

        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, inputs = tringular.make_combinations(42, seed=42, shuffle=True, reorder=False)
        >>> print(result.head(6))
           Panelist Sample 1 Sample 2 Sample 3
        0         1    B-516    B-523    A-666
        1         2    A-903    B-866    B-759
        2         3    A-953    B-390    A-944
        3         4    B-698    A-380    A-762
        4         5    A-895    A-682    B-819
        5         6    B-791    A-719    B-796
        >>> print(inputs)
        {'n_of_assessors': 42, 'seed': 42, 'shuffle': True, 'reorder': False}


        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, inputs = tringular.make_combinations(42, seed=42, shuffle=False, reorder=False)
        >>> print(result.head(6))
           Panelist Sample 1 Sample 2 Sample 3
        0         1    A-516    A-523    B-666
        1         2    A-903    B-866    A-759
        2         3    A-953    B-390    B-944
        3         4    B-698    B-380    A-762
        4         5    B-895    A-682    B-819
        5         6    B-791    A-719    A-796
        >>> print(inputs)
        {'n_of_assessors': 42, 'seed': 42, 'shuffle': False, 'reorder': False}


        >>> from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, inputs = tringular.make_combinations(42, seed=42, shuffle=False, reorder=True)
        >>> print(result.head(6))
           Panelist Sample 1 Sample 2 Sample 3
        0         1    A-174    B-548    A-462
        1         2    A-245    B-457    A-672
        2         3    A-772    B-465    B-492
        3         4    A-474    B-657    B-839
        4         5    B-895    A-682    B-819
        5         6    B-283    B-257    A-945
        >>> print(inputs)
        {'n_of_assessors': 42, 'seed': 42, 'shuffle': False, 'reorder': True}


        """

        # ----- checking n_of_assessors value ----- #
        checkers._check_is_integer(n_of_assessors, "n_of_assessors", self.language)
        checkers._check_value_is_equal_or_higher_than(n_of_assessors, "n_of_assessors", 6, language=self.language)

        # ----- checking seed value ----- #
        if seed is None:
            seed = np.random.randint(1,100)
        else:
            checkers._check_is_integer(seed, "seed", self.language)
            checkers._check_value_is_equal_or_higher_than(seed, "seed", 0, language=self.language)

        # ----- checking shuffle value ----- #
        if shuffle is None:
            shuffle = True
        else:
            checkers._check_is_bool(shuffle, "shuffle", self.language)

        # ----- checking reorder value ----- #
        if reorder is None:
            reorder = False
        else:
            checkers._check_is_bool(reorder, "reorder", self.language)


        # ----- creating the three digit random numbers ----- #
        data = np.arange(100, 1000)

        # --- Checking if the amount of unique numbers is less than the amount that will be used --- #
        # --- If not, we will replicate all 900 numbers until the correct amount is reached --- #
        while len(data) < n_of_assessors*3:
            data = np.repeat(data, 2)

        # ----- getting random sequence ----- #
        rng = np.random.default_rng(seed)
        random_sequence = rng.choice(data, n_of_assessors*3, replace=False)

        # ----- splitting the random numbers into 3 columns ----- #
        random_sequence = np.array_split(random_sequence, 3)

        # ----- all permutations ----- #
        combinations = ["AAB", "ABA", "ABB", "BBA", "BAB", "BAA"]


        # ----- if shuffle is True, we need to shuffle the list ----- #
        if shuffle:
            combinations = list(rng.choice(combinations, len(combinations), replace=False))

        # ------ Now we need to repeat the combinations to get one combination to each panelist ----- #
        all_combinations = _repeat_items(combinations, n_of_assessors)

        # ----- quering ----- #
        fk_id_function = management._query_func_id("discriminative_tests")
        messages = management._get_messages(fk_id_function, self.language, "discriminative_tests")
        panelist = messages[2][0][0]
        sample = messages[2][1][0]

        # ----- making the DataFrame ----- #
        df = pd.DataFrame(columns=[panelist, f"{sample} 1", f"{sample} 2", f"{sample} 3"])
        df[panelist] = range(1, n_of_assessors+1)
        first = []
        second = []
        third = []

        for i in range(n_of_assessors):
            first.append(f"{all_combinations[i][0]}-{random_sequence[0][i]}")
            second.append(f"{all_combinations[i][1]}-{random_sequence[1][i]}")
            third.append(f"{all_combinations[i][2]}-{random_sequence[2][i]}")

        df[f"{sample} 1"] = first
        df[f"{sample} 2"] = second
        df[f"{sample} 3"] = third

        # reording the df
        if reorder:
            df = df.sample(frac=1, random_state=seed).reset_index(drop=True)
            df[panelist] = range(1, n_of_assessors+1)

        inputs = {
            "n_of_assessors": n_of_assessors,
            "seed": seed,
            "shuffle": shuffle,
            "reorder": reorder,
        }

        self.df_table = df

        return df, inputs



    # with tests, with text, with database, with docstring
    def make_protocol(self, test_code=None, prod_name=None, sample_A=None, sample_B=None, title=None, info=None,  code_text=None, recommendations=None, file_name=None):
        """This function creates an excel protocol to prepare the Triangular test

        Parameters
        ----------
        test_code : ``str``, optional
            The code of the test
        prod_name : ``str``, optional
            The product name
        sample_A : ``str``, optional
            The name of the first sample (Sample A)
        sample_B : ``str``, optional
            The name of the second sample (Sample B)
        title : ``str``, optional
            The title of the test
        info : ``str``, optional
            Some information about the test
        code_text : ``str``, optional
            Some information about the test
        recommendations : ``list`` of ``str``
            A list with some recomendations to perform the test
        file_name : ``str``, optional
            The name of the file to export without its extension (default is ``"protocol_sheet"``)

        Notes
        -----
        The method ``make_combinations`` must be called beforehand.


        Examples
        --------

        >>> from pycafee.sensoryanalysis import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, inpupt = tringular.make_combinations(36, seed=42)
        >>> tringular.make_protocol()
        The 'protocol_sheet.xlsx' file was exported!

        .. image:: img/protocol_sheet.png
           :alt: Image showing the protocol sheet
           :align: center


        >>> from pycafee.sensoryanalysis import TriangleTest
        >>> tringular = TriangleTest(language="pt-br")
        >>> result, inpupt = tringular.make_combinations(36, seed=42, reorder=True)
        >>> tringular.make_protocol()
        O arquivo 'planilha_protocolo.xlsx'  foi exportado!

        .. image:: img/planilha_protocolo.png
           :alt: Image showing the protocol sheet
           :align: center



        """

        # ----- Cheking the parameters ---- #
        # ----- quering ----- #
        fk_id_function = management._query_func_id("discriminative_tests")
        messages = management._get_messages(fk_id_function, self.language, "discriminative_tests")

        if test_code is None:
            test_code = "000"
        else:
            checkers._check_is_str(test_code, "test_code", self.language)
        if title is None:
            title = messages[4][2][0]
        else:
            checkers._check_is_str(title, "title", self.language)
        if info is None:
            info = messages[4][3][0]
        else:
            checkers._check_is_str(info, "info", self.language)
        if prod_name is None:
            prod_name = messages[4][4][0]
        else:
            checkers._check_is_str(prod_name, "prod_name", self.language)
        if sample_A is None:
            sample_A = messages[4][5][0]
        else:
            checkers._check_is_str(sample_A, "sample_A", self.language)
        if sample_B is None:
            sample_B = messages[4][6][0]
        else:
            checkers._check_is_str(sample_B, "sample_B", self.language)
        if code_text is None:
            code_text = messages[4][7][0]
        else:
            checkers._check_is_str(code_text, "code_text", self.language)
        if recommendations is None:
            recommendations = [messages[4][8][0], messages[4][9][0], messages[4][10][0], messages[4][11][0]]
        else:
            checkers._check_is_list(recommendations, "recommendations", self.language)
            checkers._check_list_length_equal_or_higher_than(recommendations, 1, "recommendations", self.language)
            for i in range(len(recommendations)):
                checkers._check_is_str(recommendations[i], f"recommendations[{i}]", self.language)


        date = messages[4][0][0]
        text_code = messages[4][1][0]


        # batizando o nome do arquivo
        if file_name is None:
            file_name = messages[1][6][0]
        else:
            checkers._check_is_str(file_name, 'file_name', self.language)
            helpers._check_forbidden_character(file_name, "file_name", self.language)


        # ----- creating the workbook ----- #
        wb = Workbook()
        ws = wb.active
        # ------ worksheet name ----- #
        ws.title = messages[1][0][0]

        # ----- Header ----- #
        # ----- ROW 1 ----- #
        row = 1
        # --- current date --- #
        ws.merge_cells('A1:C1')
        cell = ws.cell(row=row, column=1)
        cell.value = f"{date}: {datetime.now().strftime('%Y/%m/%d')}"
        cell.font = Font(size=10, bold=True,)

        cell.alignment = Alignment(horizontal='left', vertical='center', indent=2)
        general._openpyxl_border_corner(cell, canto="superior-esquerdo")

        # --- test code --- #
        ws.merge_cells('G1:I1')
        cell = ws.cell(row=row, column=7)
        cell.value = f'{text_code}: {test_code}'
        cell.font = Font(size=10, bold=True,)
        cell.alignment = Alignment(horizontal='right', vertical='center', indent=2)

        cell = ws.cell(row=row, column=9)
        general._openpyxl_border_corner(cell, canto="superior-direito")

        # top
        for i in range(2,9):
            cell = ws.cell(row=row, column=i)
            general._openpyxl_border_corner(cell, canto="top")

        # ----- Title ----- #
        # ----- ROW 2 ----- #
        row = 2
        ws.merge_cells('A2:I2')
        cell = ws.cell(row=row, column=1)
        cell.value = title
        cell.alignment = Alignment(horizontal='center', vertical='top')
        cell.font = Font(size=12, bold=True,)
        ws.row_dimensions[row].height = 25

        cell = ws.cell(row=row, column=1)
        general._openpyxl_border_corner(cell, canto="esquerdo")

        cell = ws.cell(row=row, column=9)
        general._openpyxl_border_corner(cell, canto="direito")

        # --- ROW 3 --- #
        row = 3
        cell = ws.cell(row=row, column=1)
        cell.border = Border(
                top=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thick', color='FF000000'),
                )
        cell = ws.cell(row=row, column=9)
        cell.border = Border(
                top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thick', color='FF000000'),
                )

        for i in range(2,9):
            cell = ws.cell(row=row, column=i)
            general._openpyxl_border_corner(cell, canto="top-thin")

        # ----- INFO ----- #
        ws.merge_cells('A3:I3')
        cell = ws.cell(row=row, column=1)
        cell.value = info
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        cell.font = Font(size=11)
        ws.row_dimensions[row].height = 40

        # ----- ROW 4 ----- #
        row = 4
        cell = ws.cell(row=row, column=1)
        cell.border = Border(
                top=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thick', color='FF000000'),
                )
        cell = ws.cell(row=row, column=9)
        cell.border = Border(
                top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thick', color='FF000000'),
                )

        for i in range(2,9):
            cell = ws.cell(row=row, column=i)
            general._openpyxl_border_corner(cell, canto="top-thin")

        # ----- PRODUCT NAME ----- #
        ws.merge_cells('A4:I4')
        cell = ws.cell(row=row, column=1)
        cell.value = prod_name
        cell.font = Font(size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25

        # ----- ROW 5 ----- #
        row = 5
        # ----- SAMPLE A INFO ----- #
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"{messages[2][2][0]} A: {sample_A}"
        cell.font = Font(size=11, )
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 20

        # ----- ROW 6 ----- #
        row = 6
        # ----- SAMPLE B INFO ----- #
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = f"{messages[2][2][0]} B: {sample_B}"
        cell.font = Font(size=11, )
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 20

        # --- Border for row 5 and 6 --- #
        for i in range(5,7):
            cell = ws.cell(row=i, column=1)
            general._openpyxl_border_corner(cell, canto="esquerdo")

            cell = ws.cell(row=i, column=9)
            general._openpyxl_border_corner(cell, canto="direito")

        # ----- ROW 7 ----- #
        row = 7
        cell = ws.cell(row=row, column=1)
        cell.border = Border(
                top=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thick', color='FF000000'),
                )
        cell = ws.cell(row=row, column=9)
        cell.border = Border(
                top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thick', color='FF000000'),
                )

        for i in range(2,9):
            cell = ws.cell(row=row, column=i)
            general._openpyxl_border_corner(cell, canto="top-thin")


        # ----- adding the CODES ----- #

        ws.merge_cells(f'A{row}:I{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = code_text
        cell.font = Font(size=11, )
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 20


        # ----- ROW 8 ----- #
        row += 1
        # ----- getting the data ----- #
        if self.df_table is None:
            try:
                error = messages[5][0][0]
                raise ValueError(error)
            except ValueError:
                general._display_two_line_attention(
                    messages[6][0][0],
                    messages[6][1][0],
                                    )
                raise
        df = self.df_table.copy()
        # ----- ADDING HEADERS ----- #
        general.insert_in_cell(ws.cell(row=row, column=1), df.columns[0])
        general.insert_in_cell(ws.cell(row=row, column=2), df.columns[1])
        general.insert_in_cell(ws.cell(row=row, column=3), df.columns[2])
        general.insert_in_cell(ws.cell(row=row, column=4), df.columns[3])
        general.insert_in_cell(ws.cell(row=row, column=6), df.columns[0])
        general.insert_in_cell(ws.cell(row=row, column=7), df.columns[1])
        general.insert_in_cell(ws.cell(row=row, column=8), df.columns[2])
        general.insert_in_cell(ws.cell(row=row, column=9), df.columns[3])
        ws.row_dimensions[row].height = 20


        # --- finding the best way to split the data into two main columns --- #
        # since the test bases has 6 options, lets split the data in multiples of 6 and find the next multiple of 6 after half size
        def find_closest(value):
            for i in range(1, 7):
                if (value + i) % 6 == 0:
                    return value + i

        closest_position = find_closest(ceil(df.shape[0]/2))

        # inserting each data into the excel file
        for i in range(df.shape[0]):
            # first main colun
            if i < closest_position:
                general.insert_in_cell(ws.cell(row=row + df[df.columns[0]].iloc[i], column=1), df[df.columns[0]].iloc[i])
                general.insert_in_cell(ws.cell(row=row + df[df.columns[0]].iloc[i], column=2), df[df.columns[1]].iloc[i])
                general.insert_in_cell(ws.cell(row=row + df[df.columns[0]].iloc[i], column=3), df[df.columns[2]].iloc[i])
                general.insert_in_cell(ws.cell(row=row + df[df.columns[0]].iloc[i], column=4), df[df.columns[3]].iloc[i])

            else: # first main colun
                aux = row - closest_position #+ even_odd
                general.insert_in_cell(ws.cell(row= aux + df[df.columns[0]].iloc[i], column=6), df[df.columns[0]].iloc[i])
                general.insert_in_cell(ws.cell(row= aux + df[df.columns[0]].iloc[i], column=7), df[df.columns[1]].iloc[i])
                general.insert_in_cell(ws.cell(row= aux + df[df.columns[0]].iloc[i], column=8), df[df.columns[2]].iloc[i])
                general.insert_in_cell(ws.cell(row= aux + df[df.columns[0]].iloc[i], column=9), df[df.columns[3]].iloc[i])


            # adding space when is multiple of 6
            if (i + 1) % 6 == 0:
                ws.row_dimensions[row + i + 1].height = 20



        new_row = closest_position + row + 1

        for i in range(row, new_row):
            cell = ws.cell(row=i, column=1)
            general._openpyxl_border_corner(cell, canto="esquerdo")

            cell = ws.cell(row=i, column=9)
            general._openpyxl_border_corner(cell, canto="direito")

        row = new_row
        cell = ws.cell(row=row, column=1)
        cell.border = Border(
                top=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thick', color='FF000000'),
                )
        cell = ws.cell(row=row, column=9)
        cell.border = Border(
                top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thick', color='FF000000'),
                )

        row = new_row
        for i in range(2,9):
            cell = ws.cell(row=row, column=i)
            general._openpyxl_border_corner(cell, canto="top-thin")

        # ----- RECOMENDATIONS ----- #

        ws.merge_cells(f'A{row}:I{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = recommendations[0]
        cell.font = Font(size=12, )
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25


        for i in range(len(recommendations)):
            if i == 0:
                pass
            else:
                row += 1
                ws.merge_cells(f'A{row}:I{row}')
                cell = ws.cell(row=row, column=1)
                cell.value = recommendations[i]
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal='justify', vertical='center', indent=1)
                ws.row_dimensions[row].height = 20*ceil(int(str(len(recommendations[i])))/100)
                # ws.row_dimensions[row].height = 20

                cell = ws.cell(row=row, column=1)
                general._openpyxl_border_corner(cell, canto="esquerdo")

                cell = ws.cell(row=row, column=9)
                general._openpyxl_border_corner(cell, canto="direito")


        row += 1
        for i in range(1,10):
            cell = ws.cell(row=row, column=i)
            general._openpyxl_border_corner(cell, canto="top")

        # ----- EXPORTING THE FILE ----- #

        file = file_name + ".xlsx"
        # cheking if the file already exists. If it does, ask the user if wants to replace the file
        file_exists = helpers._check_file_exists(file)

        func_name = "_export_to_csv"
        fk_id_function = management._query_func_id(func_name)
        messages = management._get_messages(fk_id_function, self.language, func_name)

        # if the file already exists, create a new name for the file
        if file_exists:
            # warn that the file already exists
            aviso = f"{messages[1][0][0]} '{file}' {messages[1][2][0]}"
            i = 0
            # try to create a new name by adding a number to the end of the name.
            while file_exists:
                i += 1
                file = file_name + "_" + str(i) + ".xlsx"
                file_exists = helpers._check_file_exists(file)
            file_name = file
            # warn the user the name of the file to be exported
            texto = f"     ---> {messages[2][0][0]} '{file_name}'"
            general._display_warn(aviso, texto)
        else:
            file_name = file_name + '.xlsx'


        try:
            wb.save(file_name)
            general._display_one_line_success(
                text = f"{messages[3][0][0]} '{file_name}' {messages[3][2][0]}"
            )
        except PermissionError:
            # logging.error(traceback.format_exc())
            general._display_two_line_attention(
                text1 = f"{messages[4][0][0]} '{file_name}' {messages[4][2][0]}.",
                text2 = f"{messages[5][0][0]} '{file_name}' {messages[5][2][0]}"
                )
            raise
        except FileNotFoundError: # acredito que o problema com subfolder é resolvido com a proibição do /
            # logging.error(traceback.format_exc())
            general._display_two_line_attention(
                text1 = f"{messages[4][0][0]} '{file_name}' {messages[4][2][0]}.",
                text2 = f"{messages[6][0][0]}"
                )
            raise
        wb.close()


    # with tests, with text, with database, with docstring
    def make_scoresheets(self, which=None, title=None, instructions_text=None, file_name=None):
        """This function creates an excel scoresheet to perform the Triangular test

        Parameters
        ----------
        which : ``str``, optional
            Controls the number of scoresheet in each sheet.

            * If ``which="full"`` (default), all scoresheet are created in a single worksheet.
            * If ``which="two"``, multiple worksheet are created each containing two scoresheets.
            * If ``which="one"``, multiple worksheet are created each containing one scoresheet.

        title : ``str``, optional
            The title of the test. Default is ``"Triangle Test"``
        instructions_text : ``list`` of ``str``
            A list with some recomendations to perform the test. Default is ``['Evaluate samples from left to rigth. Two are alike. Mark an "X" in the box from the sample which differs from the others. If no difference is apparent, you must guess.']``.
        file_name : ``str``, optional
            The name of the file to export without its extension. Default is ``"score_sheet"``


        Notes
        -----
        The method ``make_combinations`` must be called beforehand.

        Examples
        --------

        >>> from pycafee.sensoryanalysis import TriangleTest
        >>> tringular = TriangleTest()
        >>> result, inpupt = tringular.make_combinations(36, seed=42)
        >>> tringular.make_scoresheets()
        The 'score_sheet.xlsx' file was exported!

        .. image:: img/score_sheet.png
           :alt: Image showing the protocol sheet
           :align: center


        >>> from pycafee.sensoryanalysis import TriangleTest
        >>> tringular = TriangleTest(language="pt-br")
        >>> result, inpupt = tringular.make_combinations(36, seed=42)
        >>> tringular.make_scoresheets(which="one")
        O arquivo 'planilha_score.xlsx'  foi exportado!

        .. image:: img/planilha_score.png
           :alt: Image showing the protocol sheet
           :align: center


        """

        # ----- Cheking the parameters ---- #
        # ----- checking the WHICH parameter ----- #
        if which == None or which == "full":
            which = "full"
        else:
            checkers._check_is_str(which, "which", self.language)
            if which == "two":
                which = "two"
            elif which == "one":
                which = "one"
            else:
                # ----- quering ----- #
                fk_id_function = management._query_func_id("generic")
                messages = management._get_messages(fk_id_function, self.language, "generic")
                try:
                    error = messages[3][0][0]
                    raise ValueError(error)
                except ValueError:
                    msg = [f"{messages[4][0][0]} 'which' {messages[4][2][0]}:"]
                    values = ['full', 'two', 'one']
                    for item in values:
                        msg.append(f"   --->    '{item}'")
                    msg.append(f"{messages[4][4][0]}:")
                    msg.append(f"   --->    '{which}'")
                    general._display_n_line_attention(msg)
                    raise


        # ----- quering ----- #
        fk_id_function = management._query_func_id("discriminative_tests")
        messages = management._get_messages(fk_id_function, self.language, "discriminative_tests")
        # ----- Title ----- #
        if title is None:
            title = messages[7][0][0]
        else:
            checkers._check_is_str(title, "title", self.language)
        # ----- instructions_text ----- #
        if instructions_text is None:
            instructions_text = [messages[7][3][0]]
        else:
            checkers._check_is_list(instructions_text, "instructions_text", self.language)
            checkers._check_list_length_equal_or_higher_than(instructions_text, 1, "instructions_text", self.language)
            for i in range(len(instructions_text)):
                checkers._check_is_str(instructions_text[i], f"instructions_text[{i}]", self.language)

        # ----- date, name, instructions and remarks from database ----- #
        date = messages[4][0][0]
        name = messages[7][1][0]
        instructions = messages[7][2][0]
        remarks = messages[7][4][0]

        # ----- file_name ----- #
        if file_name is None:
            file_name = messages[1][5][0]
        else:
            checkers._check_is_str(file_name, 'file_name', self.language)
            helpers._check_forbidden_character(file_name, "file_name", self.language)

        # ----- getting the data ----- #
        if self.df_table is None:
            try:
                error = messages[5][0][0]
                raise ValueError(error)
            except ValueError:
                general._display_two_line_attention(
                    messages[6][0][0],
                    messages[6][1][0],
                                    )
                raise
        df = self.df_table.copy()

        # ----- creating a new workbook ----- #
        wb = Workbook()

        # ----- first row ----- #
        first_row = 1

        # ----- create the scoresheet based on which param ----- #
        if which == "full":
            # ----- all scoresheets in one sheet ----- #
            # ----- ativando a scoresheet ----- #
            ws = wb.active
            # ------ worksheet name ----- #
            ws.title = title
            # looping for each index to generate the scoresheet
            for i in range(df.shape[0]):
                first_sample = int(df[df.columns[1]].iloc[i][-3:])
                second_sample = int(df[df.columns[2]].iloc[i][-3:])
                third_sample = int(df[df.columns[3]].iloc[i][-3:])
                panelist_name = f"{df.columns[0]} {df[df.columns[0]].iloc[i]}"
                first_row = general._openpyxl_make_scoresheets(first_row, first_sample, second_sample, third_sample, ws, panel_name=panelist_name, messages=messages, name=name, date=date, instructions=instructions, instructions_text=instructions_text, remarks=remarks, title=title)
                first_row += 3 # +3 to be nice to export to pdf

        elif which == "two":
            # aux values
            aux = 0
            count = 1
            # looping for each index to generate the scoresheet
            for i in range(df.shape[0]):
                # every two scoresheets a new sheet must be created
                if aux == 0:
                    # se aux for zero, ativamos a sheet
                    if i == 0:
                        # caso seja a primeira sheet, precisa renomear
                        ws = wb.active
                        ws.title = f"sheet_{count}"
                    else:
                        # caso contrario, basta criar nova sheet
                        ws = wb.create_sheet(f"sheet_{count}")
                    first_sample = int(df[df.columns[1]].iloc[i][-3:])
                    second_sample = int(df[df.columns[2]].iloc[i][-3:])
                    third_sample = int(df[df.columns[3]].iloc[i][-3:])
                    panelist_name = f"{df.columns[0]} {df[df.columns[0]].iloc[i]}"
                    first_row = general._openpyxl_make_scoresheets(first_row, first_sample, second_sample, third_sample, ws, panel_name=panelist_name, messages=messages, name=name, date=date, instructions=instructions, instructions_text=instructions_text, remarks=remarks, title=title)
                    first_row += 3
                    aux += 1
                else:

                    count += 1
                    aux = 0
                    first_sample = int(df[df.columns[1]].iloc[i][-3:])
                    second_sample = int(df[df.columns[2]].iloc[i][-3:])
                    third_sample = int(df[df.columns[3]].iloc[i][-3:])
                    panelist_name = f"{df.columns[0]} {df[df.columns[0]].iloc[i]}"
                    first_row = general._openpyxl_make_scoresheets(first_row, first_sample, second_sample, third_sample, ws, panel_name=panelist_name, messages=messages, name=name, date=date, instructions=instructions, instructions_text=instructions_text, remarks=remarks, title=title)
                    first_row = 1

        else:
            # looping for each index to generate the scoresheet
            for i in range(df.shape[0]):
                # we need to get the name first due to it is used as sheetname
                panelist_name = f"{df.columns[0]} {df[df.columns[0]].iloc[i]}"
                if i == 0:
                    # caso seja a primeira sheet, precisa renomear
                    ws = wb.active
                    ws.title = panelist_name
                else:
                # caso contrario, basta criar nova sheet
                    ws = wb.create_sheet(panelist_name)

                first_sample = int(df[df.columns[1]].iloc[i][-3:])
                second_sample = int(df[df.columns[2]].iloc[i][-3:])
                third_sample = int(df[df.columns[3]].iloc[i][-3:])

                first_row = general._openpyxl_make_scoresheets(first_row, first_sample, second_sample, third_sample, ws, panel_name=panelist_name, messages=messages, name=name, date=date, instructions=instructions, instructions_text=instructions_text, remarks=remarks, title=title)
                first_row = 1

        # ----- EXPORTING THE FILE ----- #

        file = file_name + ".xlsx"
        # cheking if the file already exists. If it does, ask the user if wants to replace the file
        file_exists = helpers._check_file_exists(file)

        func_name = "_export_to_csv"
        fk_id_function = management._query_func_id(func_name)
        messages = management._get_messages(fk_id_function, self.language, func_name)

        # if the file already exists, create a new name for the file
        if file_exists:
            # warn that the file already exists
            aviso = f"{messages[1][0][0]} '{file}' {messages[1][2][0]}"
            i = 0
            # try to create a new name by adding a number to the end of the name.
            while file_exists:
                i += 1
                file = file_name + "_" + str(i) + ".xlsx"
                file_exists = helpers._check_file_exists(file)
            file_name = file
            # warn the user the name of the file to be exported
            texto = f"     ---> {messages[2][0][0]} '{file_name}'"
            general._display_warn(aviso, texto)
        else:
            file_name = file_name + '.xlsx'


        try:
            wb.save(file_name)
            general._display_one_line_success(
                text = f"{messages[3][0][0]} '{file_name}' {messages[3][2][0]}"
            )
        except PermissionError:
            # logging.error(traceback.format_exc())
            general._display_two_line_attention(
                text1 = f"{messages[4][0][0]} '{file_name}' {messages[4][2][0]}.",
                text2 = f"{messages[5][0][0]} '{file_name}' {messages[5][2][0]}"
                )
            raise
        except FileNotFoundError: # acredito que o problema com subfolder é resolvido com a proibição do /
            # logging.error(traceback.format_exc())
            general._display_two_line_attention(
                text1 = f"{messages[4][0][0]} '{file_name}' {messages[4][2][0]}.",
                text2 = f"{messages[6][0][0]}"
                )
            raise


        wb.close()


    def fit(self, n_of_correct_answers, total_of_answers, alfa=None, details=None):
        """This function applies the Triangular test [1]_

        Parameters
        ----------
        n_of_correct_answers : ``int``, positive
            The number of correct responses.
        total_of_answers : ``int``, positive
            The total number of assessors.
        alfa : ``float``, optional
            The level of significance (``ɑ``). Default is ``None`` which results in ``0.05`` (``ɑ = 5%``).
        details : ``str``, optional
            The ``details`` parameter determines the amount of information presented about the hypothesis test.

            * If ``details = "short"`` (or ``None``, e.g, the default), a simplified version of the test result is returned.
            * If ``details = "full"``, a detailed version of the hypothesis test result is returned.
            * if ``details = "binary"``, the conclusion will be ``1`` (:math:`H_0` is rejected) or ``0`` (:math:`H_0` is accepted).

        Returns
        -------
        result : ``tuple`` with
            pd : ``float``
                The proportion of distinguishers
            pc : ``float``
                The proportion correct answers
            pd_std : ``float``
                The standard deviation of the proportion of the distinguishers
            pc_ic : ``float``
                The confidence interval of the proportion of the distinguishers
            ic_proportion_distinguishers : ``tuple`` with:
                lower_limit : ``float``
                    The lower confidence limit
                upper_limit : ``float``
                                        The upper confidence limit
            critical_z : ``float``
                The critical value of the Normal distribution (one-sided) corresponding to alpha;

        conclusion : ``str``
            The test conclusion (e.g, Samples are equal/not equal).

        See Also
        --------
        pycafee.sensoryanalysis.discriminative_tests.TriangleTest.make_scoresheets
        pycafee.sensoryanalysis.discriminative_tests.TriangleTest.make_protocol
        pycafee.sensoryanalysis.discriminative_tests.TriangleTest.make_combinations
        pycafee.sensoryanalysis.discriminative_tests.TriangleTest.minimum_of_correct_responses
        pycafee.sensoryanalysis.discriminative_tests.TriangleTest.get_number_assessors


        Notes
        -----
        The **Triangular Test** has the following premise:

        .. admonition:: \u2615

           :math:`H_0:` the two samples are equal.

           :math:`H_1:` the two samples are different.


        The proportion of correct answers (:math:`pc`) is estimated as follows:

        .. math::

                pc = \\frac{nc}{n}

        where :math:`nc` is the number of correct responses (``n_of_correct_answers``) and :math:`n` is the total number of assessors (``total_of_answers``).

        The proportion of judges who can differentiate the samples (:math:`pd`) is estimated as follows:

        .. math::

                pd = 1.5 pc - 0.5

        The standard deviation of the proportion  (:math:`pd_{std}`) is estimated as follows:

        .. math::

                pd_{std} = 1.5\\sqrt{\\frac{pc(1-pc)}{n}}

        The confidence interval of the proportion  (:math:`pd_{ic}`) is estimated as follows:

        .. math::

                pd_{ic} = Z \\times pd_{std}

        where :math:`Z` is the critical value of the standard Normal distribution (one-sided). This value is obtained using ``stats.norm.ppf(1-alfa)``.



        References
        ----------
        .. [1] Standard Test Method for Sensory Analysis—Triangle Test, Designation: E1885 − 04, 2011.


        Examples
        --------



        """

        # ----- checking alpha value ----- #
        if alfa is None:
            alfa = self.alfa
        else:
            # --- should be float --- #
            checkers._check_is_float(alfa, "alfa", self.language)
            # --- should be in the range (0,1) --- #
            checkers._check_data_in_range(alfa, "alfa", 0.0, 1.0, self.language)
            # --- must be in --- #
            alloweds = [0.20, 0.10, 0.05, 0.01, 0.001]
            print(alfa)
            helpers._raises_wrong_param("alfa", alloweds, alfa, self.language)
            self.alfa = alfa

        # --- cheking n_of_correct_answers --- #
        checkers._check_is_integer(n_of_correct_answers, 'n_of_correct_answers', self.language)
        checkers._check_is_positive(n_of_correct_answers, 'n_of_correct_answers', self.language)

        # --- cheking total_of_answers --- #
        checkers._check_is_integer(total_of_answers, 'total_of_answers', self.language)
        checkers._check_is_positive(total_of_answers, 'total_of_answers', self.language)

        fk_id_function = management._query_func_id("discriminative_tests")
        messages = management._get_messages(fk_id_function, self.language, "discriminative_tests")

        # --- comparing total_of_answers with n_of_correct_answers --- #
        if total_of_answers < n_of_correct_answers:
                try:
                    error = messages[7][0][0]
                    raise ValueError(error)
                except ValueError:
                    msg = [messages[7][1][0]]
                    msg.append(f"'total_of_answers = '{total_of_answers}'")
                    msg.append(f"'n_of_correct_answers = '{n_of_correct_answers}'")
                    general._display_n_line_attention(msg)
                    raise

        ### checking the details parameter ###
        if details == None:
            details = "short"
        else:
            checkers._check_is_str(details, "details", self.language)
            if details == "short":
                details = "short"
            elif details == "full":
                details = "full"
            elif details == "binary":
                details = "binary"
            else:
                fk_id_function = management._query_func_id("generic")
                messages = management._get_messages(fk_id_function, self.language, "generic")
                try:
                    error = messages[3][0][0]
                    raise ValueError(error)
                except ValueError:
                    msg = [f"{messages[4][0][0]} 'details' {messages[4][2][0]}:"]
                    values = ['short', 'full', 'binary']
                    for item in values:
                        msg.append(f"   --->    '{item}'")
                    msg.append(f"{messages[4][4][0]}:")
                    msg.append(f"   --->    '{details}'")
                    general._display_n_line_attention(msg)
                    raise


        proportion_correct = n_of_correct_answers/total_of_answers

        proportion_distinguishers = 1.5*proportion_correct - 0.5

        sd_proportion_distinguishers = 1.5*np.sqrt(proportion_correct*(1 - proportion_correct)/total_of_answers)

        critical_z = stats.norm.ppf(1- alfa)

        ic_proportion_distinguishers = sd_proportion_distinguishers*critical_z
        lower_limit = proportion_distinguishers - ic_proportion_distinguishers
        upper_limit = proportion_distinguishers + ic_proportion_distinguishers

        if lower_limit <= proportion_distinguishers <= upper_limit:
            if details == 'short':
                msg = f"{messages[8][0][0]} {100*(1-alfa)}{messages[8][2][0]}"
            elif details == 'full':
                msg = f"{messages[9][0][0]}{helpers._truncate(proportion_distinguishers, self.language, decs=self.n_digits)}{messages[9][2][0]}{helpers._truncate(lower_limit, self.language, decs=self.n_digits)}{messages[9][4][0]}{helpers._truncate(upper_limit, self.language, decs=self.n_digits)}{messages[9][6][0]} {(1-alfa)*100}{messages[9][8][0]}"
            else:
                msg = 0
        else:
            if details == 'short':
                msg = f"{messages[10][0][0]} {100*(1-alfa)}{messages[10][2][0]}"
            elif details == 'full':
                if proportion_distinguishers < lower_limit:
                    msg = f"{messages[11][0][0]}{helpers._truncate(proportion_distinguishers, self.language, decs=self.n_digits)}{messages[11][2][0]}{helpers._truncate(lower_limit, self.language, decs=self.n_digits)}{messages[11][4][0]} {(1-alfa)*100}{messages[11][6][0]}"
                else:
                    msg = f"{messages[12][0][0]}{helpers._truncate(proportion_distinguishers, self.language, decs=self.n_digits)}{messages[12][2][0]}{helpers._truncate(upper_limit, self.language, decs=self.n_digits)}{messages[12][4][0]} {(1-alfa)*100}{messages[12][6][0]}"
            else:
                msg = 1

        result = namedtuple(messages[13][0][0], (messages[13][1][0], messages[13][2][0], messages[13][3][0], messages[13][4][0],
                            messages[13][5][0], messages[13][7][0]))
        self.conclusion = msg

        return result(proportion_distinguishers, proportion_correct, sd_proportion_distinguishers,
                    ic_proportion_distinguishers, (lower_limit, upper_limit), critical_z), self.conclusion




    # with tests, with text, with database (Dixon), with docstring
    def __str__(self):
        if self.conclusion is None:
            fk_id_function = management._query_func_id("generic")
            messages = management._get_messages(fk_id_function, self.language, "generic")
            return f"{messages[2][0][0]} TriangleTest {messages[2][2][0]}"
        else:
            return self.conclusion

    # with tests, with text, with database, with docstring
    def __repr__(self):
        fk_id_function = management._query_func_id("TriangularTest")
        messages = management._get_messages(fk_id_function, self.language, "TriangularTest")
        return messages[1][0][0]






#















# with some tests
def _repeat_items(lista, elemts):
    """This function returns a repetition of ``elemts`` elements of a list ``list``.

    Parameters
    ----------
    lista: ``list``
        The list that will be repeated
    elemts: ``int``
        The number of list elements with repeated items

    Returns
    -------
    lista: ``list``
        A list of repeated elements


    Examples
    --------

    >>> repeat_items([1,2], 3)
    [1, 2, 1]

    >>> repeat_items([1,2], 6)
    [1, 2, 1, 2, 1, 2]

    >>> repeat_items(["A", "B", "C", "D"], 12)
    ['A', 'B', 'C', 'D', 'A', 'B', 'C', 'D', 'A', 'B', 'C', 'D']

    >>> repeat_items(["A", "B", "C", "D"], 2)
    ['A', 'B']

    """
    return lista * (elemts // len(lista)) + lista[:(elemts % len(lista))]










#
