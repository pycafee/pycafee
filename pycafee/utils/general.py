"""This module concentrates all generic helper functions, e.g, functions that does not require language management

"""
# Function list:
#     - _display_one_line_success(text)
#     - _display_n_line_warn(flat_list)
#     - _display_one_line_attention(text)
#     - _display_n_line_attention(flat_list)
#     - _display_two_line_attention(text1, text2)
#     - _display_warn(aviso, texto)
#     - _flatten_list_of_list_string(list_of_lists)

#########################################
################ Imports ################
#########################################

###### Standard ######
from math import ceil
###### Third part ######
from colorama import Fore, Style, Back, init
from openpyxl.styles import Border, Side, Alignment, Font

###### Home made ######


###########################################
################ Functions ################
###########################################


# ----- functions to add data ----- #
def _openpyxl_insert_data(row, column, value, ws, font=11, bold=False, a_horizontal="center", a_vertical="center", a_indent=0, a_wrapText=True):
    cell = ws.cell(row=row, column=column)
    cell.value = value
    cell.alignment = Alignment(horizontal=a_horizontal, vertical=a_vertical, indent=a_indent, wrapText=a_wrapText)
    cell.font = Font(size=font, bold=bold)


def _openpyxl_border_left_right_only(cell, row, ws):
    cell = ws.cell(row=row, column=1)
    _openpyxl_border_corner(cell, canto="esquerdo")
    cell = ws.cell(row=row, column=9)
    _openpyxl_border_corner(cell, canto="direito")


def _openpyxl_make_scoresheets(first_row, first_sample, second_sample, third_sample, ws, panel_name, messages, name, date, instructions, instructions_text, remarks, title):
    # ----- Title ----- #
    # ----- ROW 1 ----- #
    row = first_row
    ws.merge_cells(f'A{row}:I{row}')
    # insert_data(row, 1, messages[1][0][0], font=14, bold=True, a_horizontal="center", a_vertical="center")
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = Font(size=14, bold=True,)
    ws.row_dimensions[row].height = 40

    # top borders
    _openpyxl_border_corner(cell, canto="superior-esquerdo")
    cell = ws.cell(row=row, column=9)
    _openpyxl_border_corner(cell, canto="superior-direito")
    for i in range(2,9):
        cell = ws.cell(row=row, column=i)
        _openpyxl_border_corner(cell, canto="top")

    # ----- INFO ----- #
    # ----- ROW 2 ----- #
    row += 1
    # borders
    cell = ws.cell(row=row, column=1)
    cell.border = Border(
            top=Side(border_style='thin', color='FF000000'),
            left=Side(border_style='thick', color='FF000000'),
            )
    cell = ws.cell(row=row, column=9)
    cell.border = Border(
            top=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thick', color='FF000000'),
            )
    for i in range(2,9):
        cell = ws.cell(row=row, column=i)
        _openpyxl_border_corner(cell, canto="top-thin")

    # data
    ws.merge_cells(f'A{row}:I{row}')
    count = 85 - (len(panel_name) + len(name) + len(date) + len(",::") + 23)
    value = f" {panel_name},    {name}: {'_'*count}  {date}: {'_'*15} "
    _openpyxl_insert_data(row, 1, value, ws=ws, font=11, a_horizontal="center", a_vertical="bottom")
    ws.row_dimensions[row].height = 30

    # empty row
    row += 1
    _openpyxl_border_left_right_only(cell, row, ws)
    ws.row_dimensions[row].height = 10

    # ----- Instructions ----- #
    # borders
    row += 1
    cell = ws.cell(row=row, column=1)
    cell.border = Border(
            top=Side(border_style='thin', color='FF000000'),
            left=Side(border_style='thick', color='FF000000'),
            )
    cell = ws.cell(row=row, column=9)
    cell.border = Border(
            top=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thick', color='FF000000'),
            )

    for i in range(2,9):
        cell = ws.cell(row=row, column=i)
        _openpyxl_border_corner(cell, canto="top-thin")

    # data
    ws.merge_cells(f'A{row}:I{row}')
    _openpyxl_insert_data(row, 1, instructions, ws=ws, font=11, bold=True, a_horizontal="left", a_vertical="bottom", a_indent=2)
    ws.row_dimensions[row].height = 25

    # ----- Instructions Text ----- #
    for instruc in instructions_text:
        # border
        row += 1
        _openpyxl_border_left_right_only(cell, row, ws)
        # data
        ws.merge_cells(f'A{row}:I{row}')
        _openpyxl_insert_data(row, 1, instruc, ws=ws, font=10, bold=False, a_horizontal="justify", a_vertical="center", a_indent=2)
        ws.row_dimensions[row].height = 20*ceil(int(str(len(instruc))[:-1])/10)


    # ----- ADDING THE RANDOM 3 DIGITs NUMBERS ----- #
    # border
    row += 1
    _openpyxl_border_left_right_only(cell, row, ws)
    # data
    numbers = [
            first_sample,
            second_sample,
            third_sample,
            ]
    columns = [3, 5, 7]
    for numb, col in zip(numbers, columns):
        _openpyxl_insert_data(row, col, numb, ws=ws, font=16, bold=True, a_horizontal="center", a_vertical="bottom")
    ws.row_dimensions[row].height = 30

    # ----- ADDING SQUARES ----- #
    # borders
    row += 1
    _openpyxl_border_left_right_only(cell, row, ws)
    # data
    columns = [3, 5, 7]
    for col in columns:
        _openpyxl_insert_data(row, col, u'\u25a1', ws=ws, font=36, bold=False, a_horizontal="center", a_vertical="bottom")
    ws.row_dimensions[row].height = 30

    # empty row
    row += 1
    _openpyxl_border_left_right_only(cell, row, ws)

    # ----- Remarks ----- #
    # borders
    row += 1
    _openpyxl_border_left_right_only(cell, row, ws)
    # data
    ws.merge_cells(f'A{row}:I{row}')
    count = 85 - (len(remarks) + 10)
    value = f"  {remarks}:   {'_'*count} "
    _openpyxl_insert_data(row, 1, value, ws=ws, font=11, bold=False, a_horizontal="center", a_vertical="center")
    ws.row_dimensions[row].height = 25

    # ----- LINES ----- #
    for i in range(2):
        row += 1
        _openpyxl_border_left_right_only(cell, row, ws)
        ws.merge_cells(f'A{row}:I{row}')
        value = f" {'_'*77} "
        _openpyxl_insert_data(row, 1, value, ws=ws, font=11, bold=False, a_horizontal="center", a_vertical="center")
        ws.row_dimensions[row].height = 25


    # ----- last line ----- #
    # border
    row += 1
    _openpyxl_border_left_right_only(cell, row, ws)
    ws.row_dimensions[row].height = 10

    row += 1
    for i in range(1,10):
        cell = ws.cell(row=row, column=i)
        _openpyxl_border_corner(cell, canto="top")
    ws.row_dimensions[row].height = 10

    return row




def insert_in_cell(cell, data):
    """This functions helps to append data into Excel forms

    Parameters
    ----------
    cell: instância de celula
    data: string or number
        The data to be inserted
    """
    cell.value = data
    cell.font = Font(size=11)
    cell.alignment = Alignment(horizontal='center', vertical='top')


def _openpyxl_border_corner(cell, canto, border_style="thick", color="FF000000"):
    """This function helps to create forms in Excel

    Parameters
    ----------
    cell: instância de celula

    canto: ``str``
        posição da borda

    """

    if canto == "superior-direito":
        cell.border = Border(
                top=Side(border_style=border_style, color=color),
                right=Side(border_style=border_style, color=color),
                )
    elif canto == "superior-esquerdo":
        cell.border = Border(
                top=Side(border_style=border_style, color=color),
                left=Side(border_style=border_style, color=color),
                )
    elif canto == "inferior-direito":
        cell.border = Border(
                bottom=Side(border_style=border_style, color=color),
                right=Side(border_style=border_style, color=color),
                )
    elif canto == "inferior-esquerdo":
        cell.border = Border(
                bottom=Side(border_style=border_style, color=color),
                left=Side(border_style=border_style, color=color),
                )
    elif canto == "top":
        cell.border = Border(
                top=Side(border_style='thick', color=color),
                )
    elif canto == "top-thin":
        cell.border = Border(
                top=Side(border_style='thin', color=color),
                )
    elif canto == "direito":
        cell.border = Border(
                right=Side(border_style='thick', color=color),
                )
    elif canto == "esquerdo":
        cell.border = Border(
                left=Side(border_style='thick', color=color),
                )
    else:
        raise ValueError



def _display_one_line_success(text):
    """
    This function prints text to the console in a simple way, but with some editing

    Parameters
    ----------
    text: ``str``
        The text to be printed

    """
    init()
    print(Fore.BLUE, "    " + text)
    print(Style.RESET_ALL)


def _display_n_line_warn(flat_list):
    """This function prints in the console the elements passed as a parameter, the first element being printed in red

    Parameters
    ----------
    flat_list: ``list``
        A flat list with values to be printed.

    """
    init()
    size = len(max(flat_list, key=len))
    print("\n")
    print(Fore.RED, "-"*size)
    i = 0
    for text in flat_list:
        if i == 0:
            print(Fore.RED, text)
        else:
            print(Fore.WHITE, text)
        i += 1
    print(Fore.RED, "-"*size)
    print("\n")
    print(Style.RESET_ALL)

def _display_one_line_attention(text):
    """
    This function prints edited text to the console in a way that catches the user's attention

    Parameters
    ----------
    text: ``str``
        The text to be printed

    """
    print("\n")
    print("^"*len(text))
    print(text)
    print("^"*len(text))
    print("\n")

def _display_n_line_attention(flat_list):
    """This function prints in the console the elements passed as a parameter, in a way that catches the user's attention

    Parameters
    ----------
    flat_list: ``list``
        A flat list with values to be printed.

    """

    size = len(max(flat_list, key=len))
    print("\n")
    print("^"*size)
    for text in flat_list:
        print(text)
    print("^"*size)
    print("\n")

def _display_two_line_attention(text1, text2):
    """
    This function prints edited text to the console in a way that catches the user's attention, where text 1 is printed on the first line, and text2 is printed on the second line.

    Parameters
    ----------
    text1: ``str``
        The text to be printed on the first line
    text2: ``str``
        The text to be printed on the second line

    """
    size = max(len(text1), len(text2))
    print("\n")
    print("^"*size)
    print(text1)
    print(text2)
    print("^"*size)
    print("\n")

def _display_warn(aviso, texto):
    """This function prints in the console the ``aviso`` in red at first line and the texto in white at second line.

    Parameters
    ----------
    aviso: ``str``
        The text to be printed in red
    texto: ``str``
        The text to be printed in white

    """
    init()
    size = max(len(aviso), len(texto))
    # print("\n")
    print(Fore.RED, "-"*size)
    print(Fore.RED, aviso)
    print(Fore.WHITE, texto)
    print(Fore.RED, "-"*size)
    # print("\n")
    print(Style.RESET_ALL)

def _flatten_list_of_list_string(list_of_lists):
    """
    This function flats a nested list with strings in it

    Parameters
    ----------
    list_of_lists: list of lists
        The list to be flattened

    Example
    -------
    a = ['The language enn is not available yet.', 'The languages available are:', ['  --->  en', '  --->  pt-br']]
    flat_list = _flatten_list_of_list_string(a)
    flat_list = list(flat_list)

    Reference
    ---------
    .. [1] https://stackoverflow.com/a/5286571/17872198
    """
    for x in list_of_lists:
        if hasattr(x, '__iter__') and not isinstance(x, str):
            for y in _flatten_list_of_list_string(x):
                yield y
        else:
            yield x











# But I remember... everything   https://youtu.be/76iQNS9VLkc?t=70
