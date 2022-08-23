"""Tests if the make_protocol is working as expected

---> Class Test_make_protocol



--------------------------------------------------------------------------------
Command to run at the prompt:
    python -m unittest -v tests/sensoryanalysis/discriminative_tests/TriangleTest/test_make_protocol.py
    or
    python -m unittest -b tests/sensoryanalysis/discriminative_tests/TriangleTest/test_make_protocol.py

--------------------------------------------------------------------------------
"""

import os
import unittest
from pycafee.utils.helpers import _check_file_exists
from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
from pathlib import Path
import pandas as pd
import io
import sys
from openpyxl import load_workbook
os.system('cls')



class Test_make_protocol_values(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.which_value = "full"
        cls.file_name_en = "protocol_sheet.xlsx"
        cls.file_name_pt = "planilha_protocolo.xlsx"



    def test_n_of_assessors(self):
        test = TriangleTest()
        test.make_combinations(20, seed=42)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, 1, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "B-751", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "B-651", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "A-814", msg="wrong 3digit value")

        row = 14
        self.assertEqual(ws[f'A{row}'].value, 6, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "B-792", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "A-245", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "A-298", msg="wrong 3digit value")

        row = 18
        self.assertEqual(ws[f'A{row}'].value, 10, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "A-530", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "A-750", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "B-655", msg="wrong 3digit value")

        row = 18
        self.assertEqual(ws[f'F{row}'].value, None, msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, None, msg="wrong 3digit value")

        rem_file.unlink()




        test = TriangleTest()
        test.make_combinations(48, seed=42)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, 1, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "A-162", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "A-845", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "B-709", msg="wrong 3digit value")

        row = 14
        self.assertEqual(ws[f'A{row}'].value, 6, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "B-478", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "A-225", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "A-773", msg="wrong 3digit value")

        row = 18
        self.assertEqual(ws[f'A{row}'].value, 10, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "A-674", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "B-415", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "B-582", msg="wrong 3digit value")

        row = 28
        self.assertEqual(ws[f'F{row}'].value, None, msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, None, msg="wrong 3digit value")

        rem_file.unlink()


        test = TriangleTest()
        test.make_combinations(21, seed=42)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, 1, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "B-712", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "A-890", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "A-273", msg="wrong 3digit value")

        row = 14
        self.assertEqual(ws[f'A{row}'].value, 6, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "B-649", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "B-653", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "A-811", msg="wrong 3digit value")

        row = 17
        self.assertEqual(ws[f'F{row}'].value, 21, msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, "A-790", msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, "B-588", msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, "A-997", msg="wrong 3digit value")

        row = 19
        self.assertEqual(ws[f'F{row}'].value, None, msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, None, msg="wrong 3digit value")

        rem_file.unlink()





    def test_seed_42(self):
        test = TriangleTest()
        test.make_combinations(37, seed=42)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)


        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, 1, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "A-817", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "B-493", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "A-480", msg="wrong 3digit value")

        row = 14
        self.assertEqual(ws[f'A{row}'].value, 6, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "B-170", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "A-562", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "B-248", msg="wrong 3digit value")

        row = 18
        self.assertEqual(ws[f'A{row}'].value, 10, msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, "B-732", msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, "A-400", msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, "A-463", msg="wrong 3digit value")

        row = 21
        self.assertEqual(ws[f'F{row}'].value, 37, msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, "A-513", msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, "B-799", msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, "A-185", msg="wrong 3digit value")

        rem_file.unlink()



    def test_default(self):
        test = TriangleTest()
        result, inputs = test.make_combinations(37)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, result[result.columns[0]].iloc[0], msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, result[result.columns[1]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, result[result.columns[2]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, result[result.columns[3]].iloc[0], msg="wrong 3digit value")

        row = 21
        self.assertEqual(ws[f'F{row}'].value, result[result.columns[0]].iloc[-1], msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, result[result.columns[1]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, result[result.columns[2]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, result[result.columns[3]].iloc[-1], msg="wrong 3digit value")


        rem_file.unlink()



    def test_shuffle(self):
        test = TriangleTest()
        result, inputs = test.make_combinations(37, shuffle=True)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, result[result.columns[0]].iloc[0], msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, result[result.columns[1]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, result[result.columns[2]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, result[result.columns[3]].iloc[0], msg="wrong 3digit value")

        row = 21
        self.assertEqual(ws[f'F{row}'].value, result[result.columns[0]].iloc[-1], msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, result[result.columns[1]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, result[result.columns[2]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, result[result.columns[3]].iloc[-1], msg="wrong 3digit value")


        rem_file.unlink()


        test = TriangleTest()
        result, inputs = test.make_combinations(37, shuffle=False)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, result[result.columns[0]].iloc[0], msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, result[result.columns[1]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, result[result.columns[2]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, result[result.columns[3]].iloc[0], msg="wrong 3digit value")

        row = 21
        self.assertEqual(ws[f'F{row}'].value, result[result.columns[0]].iloc[-1], msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, result[result.columns[1]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, result[result.columns[2]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, result[result.columns[3]].iloc[-1], msg="wrong 3digit value")


        rem_file.unlink()



    def test_reorder(self):
        test = TriangleTest()
        result, inputs = test.make_combinations(37, reorder=True)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, result[result.columns[0]].iloc[0], msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, result[result.columns[1]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, result[result.columns[2]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, result[result.columns[3]].iloc[0], msg="wrong 3digit value")

        row = 21
        self.assertEqual(ws[f'F{row}'].value, result[result.columns[0]].iloc[-1], msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, result[result.columns[1]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, result[result.columns[2]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, result[result.columns[3]].iloc[-1], msg="wrong 3digit value")


        rem_file.unlink()


        test = TriangleTest()
        result, inputs = test.make_combinations(37, reorder=False)
        test.make_protocol()
        file_name = self.file_name_en
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "TriangleTest", msg="wrong sheet name")
        ws = wb[sheets[0]]
        row = 9
        self.assertEqual(ws[f'A{row}'].value, result[result.columns[0]].iloc[0], msg="wrong panelist")
        self.assertEqual(ws[f'B{row}'].value, result[result.columns[1]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'C{row}'].value, result[result.columns[2]].iloc[0], msg="wrong 3digit value")
        self.assertEqual(ws[f'D{row}'].value, result[result.columns[3]].iloc[0], msg="wrong 3digit value")

        row = 21
        self.assertEqual(ws[f'F{row}'].value, result[result.columns[0]].iloc[-1], msg="wrong panelist")
        self.assertEqual(ws[f'G{row}'].value, result[result.columns[1]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'H{row}'].value, result[result.columns[2]].iloc[-1], msg="wrong 3digit value")
        self.assertEqual(ws[f'I{row}'].value, result[result.columns[3]].iloc[-1], msg="wrong 3digit value")


        rem_file.unlink()






class Test_make_protocol(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        # criando um arquivo xlsx generico
        file_name = "bawitdaba.xlsx"
        # file_path = folder_name + "/" + file_name
        df = pd.DataFrame(columns=["a"], data=["c"])
        df.to_excel(file_name, index=False, engine="openpyxl")

        # fazendo o teste
        result = _check_file_exists(file_name)

        # para deletar o arquivo generico
        cls.rem_file = Path(file_name)


    @classmethod
    def tearDownClass(cls):
        cls.rem_file.unlink()


    def test_when_file_exists(self):
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(file_name="bawitdaba")
        result = _check_file_exists("bawitdaba.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("bawitdaba_1.xlsx")
        rem_file.unlink()

        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol(file_name="bawitdaba")
        result = _check_file_exists("bawitdaba.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("bawitdaba_1.xlsx")
        rem_file.unlink()


    def test_when_file_exists_output(self):
        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(file_name="bawitdaba")
        rem_file = Path("bawitdaba_1.xlsx")
        rem_file.unlink()
        sys.stdout = sys.__stdout__
        expected = "UserWarning"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        expected = "'bawitdaba_1.xlsx' file"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        #################


        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol(file_name="bawitdaba")
        rem_file = Path("bawitdaba_1.xlsx")
        rem_file.unlink()
        sys.stdout = sys.__stdout__
        expected = "UserWarning: O arquivo"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        expected = "quivo 'bawitdaba_1.xlsx'  f"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        #################


    def test_pass(self):
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists("protocol_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("protocol_sheet.xlsx")
        rem_file.unlink()

        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists("planilha_protocolo.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("planilha_protocolo.xlsx")
        rem_file.unlink()


    def test_pass_output(self):
        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        sys.stdout = sys.__stdout__
        expected = "The 'protocol_sheet.xlsx' file was expo"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")
        rem_file = Path("protocol_sheet.xlsx")
        rem_file.unlink()


        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        sys.stdout = sys.__stdout__
        expected = "rquivo 'planilha_protocolo.xlsx'  foi expor"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")
        rem_file = Path("planilha_protocolo.xlsx")
        rem_file.unlink()



    def test_raises_combinations_not_generated_yet(self):
        with self.assertRaises(ValueError, msg="Does not raised error when the combinatios were not generated yet"):
            test = TriangleTest()
            test.make_protocol()


    def test_raises_combinations_not_generated_yet_output(self):

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        try:
            test = TriangleTest()
            test.make_protocol()
        except ValueError:
            pass
        sys.stdout = sys.__stdout__
        expected = "the protocol it is necessary to generate the"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        ################

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        try:
            test = TriangleTest(language="pt-br")
            test.make_protocol()
        except ValueError:
            pass
        sys.stdout = sys.__stdout__
        expected = "ilize o método 'make_combinations' para obter as "
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")


    def test_code(self):
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(test_code="001")
        result = _check_file_exists("protocol_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("protocol_sheet.xlsx")

        wb = load_workbook("protocol_sheet.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['G1'].value, "Test code: 001", msg="wrong code")
        rem_file.unlink()



        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists("planilha_protocolo.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("planilha_protocolo.xlsx")

        wb = load_workbook("planilha_protocolo.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['G1'].value, "Código do teste: 000", msg="wrong code")
        rem_file.unlink()



        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists("protocol_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("protocol_sheet.xlsx")

        wb = load_workbook("protocol_sheet.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['G1'].value, "Test code: 000", msg="wrong code")
        rem_file.unlink()


    def test_date(self):
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists("protocol_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("protocol_sheet.xlsx")

        wb = load_workbook("protocol_sheet.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        result = False
        if "Date" in ws['A1'].value:
            result = True
        self.assertTrue(result, msg="wrong date")
        rem_file.unlink()


        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists("planilha_protocolo.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("planilha_protocolo.xlsx")

        wb = load_workbook("planilha_protocolo.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        result = False
        if "Data" in ws['A1'].value:
            result = True
        self.assertTrue(result, msg="wrong date")
        rem_file.unlink()


    def test_info(self):
        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(info="BABYMETAL RULES")
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A3'].value, "BABYMETAL RULES", msg="wrong info")
        rem_file.unlink()


        file_name = "planilha_protocolo.xlsx"
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A3'].value, "Coloque esta folha na área onde as bandejas são preparadas. \nCodifique as planilhas e os contêineres com antecedência.", msg="wrong info")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A3'].value, "Post this sheet in the area where trays are prepared. \nCode scoresheets and serving containers ahead of time.", msg="wrong info")
        rem_file.unlink()


    def test_sample_A(self):
        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(sample_A="BABYMETAL RULES")
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A5'].value, "Product A: BABYMETAL RULES", msg="wrong sample_A")
        rem_file.unlink()


        file_name = "planilha_protocolo.xlsx"
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A5'].value, "Produto A: INFORMAÇÃO DA AMOSTRA A", msg="wrong sample_A")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A5'].value, "Product A: SAMPLE A INFORMATION", msg="wrong sample_A")
        rem_file.unlink()


    def test_sample_B(self):
        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(sample_B="BABYMETAL RULES")
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A6'].value, "Product B: BABYMETAL RULES", msg="wrong sample_B")
        rem_file.unlink()


        file_name = "planilha_protocolo.xlsx"
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A6'].value, "Produto B: INFORMAÇÃO DA AMOSTRA B", msg="wrong sample_B")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A6'].value, "Product B: SAMPLE B INFORMATION", msg="wrong sample_B")
        rem_file.unlink()


    def test_prod_name(self):
        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(prod_name="BABYMETAL RULES")
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A4'].value, "BABYMETAL RULES", msg="wrong prod_name")
        rem_file.unlink()


        file_name = "planilha_protocolo.xlsx"
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A4'].value, "NOME DO PRODUTO", msg="wrong prod_name")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A4'].value, "PRODUCT NAME", msg="wrong title")
        rem_file.unlink()


    def test_title(self):
        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(title="BABYMETAL RULES")
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A2'].value, "BABYMETAL RULES", msg="wrong title")
        rem_file.unlink()


        file_name = "planilha_protocolo.xlsx"
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A2'].value, "Teste Triangular - Protocolo de serviço e ordenação", msg="wrong title")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A2'].value, "Triangle test - Order and Serving Protocol", msg="wrong title")
        rem_file.unlink()


    def test_code_text(self):
        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(code_text="BABYMETAL RULES")
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A7'].value, "BABYMETAL RULES", msg="wrong code text")
        rem_file.unlink()


        file_name = "planilha_protocolo.xlsx"
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A7'].value, "Codifique os recipientes da seguinte forma:", msg="wrong code text")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A7'].value, "Code serving containers as follows:", msg="wrong code text")
        rem_file.unlink()


    def test_recommendations(self):
        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A33'].value, "RECOMMENDATIONS", msg="wrong recommendations[0]")
        self.assertEqual(ws['A34'].value, "1. Label plates with the unique 3 digit random numbers and arrange in serving order by panelist.", msg="wrong recommendations[1]")
        self.assertEqual(ws['A35'].value, "2. To serve, place samples and a coded scoresheet on a serving tray.", msg="wrong recommendations[2]")
        self.assertEqual(ws['A36'].value, "3. Decode whether reply was correct or incorrect by referring back to the worksheet.", msg="wrong recommendations[3]")
        rem_file.unlink()



        file_name = "planilha_protocolo.xlsx"
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_protocol()
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A33'].value, "RECOMENDAÇÕES", msg="wrong recommendations[0]")
        self.assertEqual(ws['A34'].value, "1. Rotule as placas com os números aleatórios exclusivos de 3 dígitos e organize-os em ordem de servir pelo provador", msg="wrong recommendations[1]")
        self.assertEqual(ws['A35'].value, "2. Para servir, coloque as amostras e uma planilha codificada em uma bandeja de servir.", msg="wrong recommendations[2]")
        self.assertEqual(ws['A36'].value, "3. Decodifique se a resposta foi correta ou incorreta consultando a planilha.", msg="wrong recommendations[3]")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(recommendations=["Baby metal", "Anderson", "Metal", "Unluck morfeus"])
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A33'].value, "Baby metal", msg="wrong recomendation")
        self.assertEqual(ws['A34'].value, "Anderson", msg="wrong recomendation")
        self.assertEqual(ws['A35'].value, "Metal", msg="wrong recomendation")
        self.assertEqual(ws['A36'].value, "Unluck morfeus", msg="wrong recomendation")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(recommendations=["Baby metal",])
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A33'].value, "Baby metal", msg="wrong recomendation")
        self.assertEqual(ws['A34'].value, None, msg="wrong recomendation")
        self.assertEqual(ws['A35'].value, None, msg="wrong recomendation")
        self.assertEqual(ws['A36'].value, None, msg="wrong recomendation")
        rem_file.unlink()


        file_name = "protocol_sheet.xlsx"
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_protocol(recommendations=["Baby metal", "Anderson", "Metal", "Unluck morfeus", "harry potter", "novalgina"])
        result = _check_file_exists(file_name)
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A33'].value, "Baby metal", msg="wrong recomendation")
        self.assertEqual(ws['A34'].value, "Anderson", msg="wrong recomendation")
        self.assertEqual(ws['A35'].value, "Metal", msg="wrong recomendation")
        self.assertEqual(ws['A36'].value, "Unluck morfeus", msg="wrong recomendation")
        self.assertEqual(ws['A37'].value, "harry potter", msg="wrong recomendation")
        self.assertEqual(ws['A38'].value, "novalgina", msg="wrong recomendation")
        rem_file.unlink()



# never an honest word, but thats when I run the world https://youtu.be/uNaHzwkDOIk?t=176

if __name__ == "__main__":
    unittest.main()
