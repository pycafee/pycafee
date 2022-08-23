"""Tests if the make_scoresheets is working as expected

---> Class Test_make_scoresheets

---> Test_make_scoresheets_which_one

---> Test_make_scoresheets_which_two

---> Test_make_scoresheets_which_full

--------------------------------------------------------------------------------
Command to run at the prompt:
    python -m unittest -v tests/sensoryanalysis/discriminative_tests/TriangleTest/test_make_scoresheets.py
    or
    python -m unittest -b tests/sensoryanalysis/discriminative_tests/TriangleTest/test_make_scoresheets.py

--------------------------------------------------------------------------------
"""

import os
import unittest
from pycafee.utils.helpers import _check_file_exists
from pycafee.sensoryanalysis.discriminative_tests import TriangleTest
from pathlib import Path
import pandas as pd
import io
import sys
from openpyxl import load_workbook


os.system('cls')



class Test_make_scoresheets_which_one(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.which_value = "one"



    def test_which_two_seed_42(self):
        test = TriangleTest()
        test.make_combinations(37, seed=42)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)


        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(sheets[0], "Panelist 1", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws['C6'].value, 817, msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, 493, msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, 480, msg="wrong 3digit value")

        result = False
        if "Panelist 1" in ws['A2'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")

        self.assertEqual(sheets[1], "Panelist 2", msg="wrong sheet name")
        ws = wb[sheets[1]]

        self.assertEqual(ws['C6'].value, 137, msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, 772, msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, 650, msg="wrong 3digit value")


        result = False
        if "Panelist 2" in ws['A2'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")


        self.assertEqual(sheets[6], "Panelist 7", msg="wrong sheet name")
        ws = wb[sheets[6]]

        self.assertEqual(ws['C6'].value, 980, msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, 623, msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, 203, msg="wrong 3digit value")

        self.assertEqual(ws['C21'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws['E21'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws['G21'].value, None, msg="wrong 3digit value")

        result = False
        if "Panelist 7" in ws['A2'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")

        self.assertEqual(ws['A17'].value, None, msg="not none")

        rem_file.unlink()


    def test_which_full_ramdom(self):
        test = TriangleTest()
        result, inputs = test.make_combinations(37)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(sheets[0], "Panelist 1", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws['C6'].value, int(result[result.columns[1]].iloc[0][-3:]), msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, int(result[result.columns[2]].iloc[0][-3:]), msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, int(result[result.columns[3]].iloc[0][-3:]), msg="wrong 3digit value")

        rem_file.unlink()



    def test_which_full(self):
        test = TriangleTest()
        test.make_combinations(37)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 37, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "Panelist 1", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws.max_row, 13, msg="row size does not match")
        self.assertEqual(ws['A1'].value, "Triangle Test", msg="wrong title")
        self.assertIsInstance(ws['C6'].value, int, msg="tri-digit not a number")
        self.assertIsInstance(ws['E6'].value, int, msg="tri-digit not a number")
        self.assertIsInstance(ws['G6'].value, int, msg="tri-digit not a number")
        rem_file.unlink()



    def test_which_full_output(self):
        test = TriangleTest()
        test.make_combinations(37)

        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput

        test.make_scoresheets(which=self.which_value)

        sys.stdout = sys.__stdout__
        expected = "he 'score_sheet.xlsx' file wa"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")
        rem_file.unlink()
        ###############



class Test_make_scoresheets_which_two(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.which_value = "two"



    def test_which_two_seed_42(self):
        test = TriangleTest()
        test.make_combinations(37, seed=42)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)


        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(sheets[0], "sheet_1", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws['C6'].value, 817, msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, 493, msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, 480, msg="wrong 3digit value")

        self.assertEqual(ws['C21'].value, 137, msg="wrong 3digit value")
        self.assertEqual(ws['E21'].value, 772, msg="wrong 3digit value")
        self.assertEqual(ws['G21'].value, 650, msg="wrong 3digit value")

        result = False
        if "Panelist 1" in ws['A2'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")

        result = False
        if "Panelist 2" in ws['A17'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")



        self.assertEqual(sheets[6], "sheet_7", msg="wrong sheet name")
        ws = wb[sheets[6]]
        self.assertEqual(ws['C6'].value, 546, msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, 329, msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, 428, msg="wrong 3digit value")

        self.assertEqual(ws['C21'].value, 804, msg="wrong 3digit value")
        self.assertEqual(ws['E21'].value, 893, msg="wrong 3digit value")
        self.assertEqual(ws['G21'].value, 503, msg="wrong 3digit value")

        result = False
        if "Panelist 13" in ws['A2'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")

        result = False
        if "Panelist 14" in ws['A17'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")



        self.assertEqual(sheets[-1], "sheet_19", msg="wrong sheet name")
        ws = wb[sheets[-1]]
        self.assertEqual(ws['C6'].value, 513, msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, 799, msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, 185, msg="wrong 3digit value")


        self.assertEqual(ws['C21'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws['E21'].value, None, msg="wrong 3digit value")
        self.assertEqual(ws['G21'].value, None, msg="wrong 3digit value")

        result = False
        if "Panelist 37" in ws['A2'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")
        self.assertEqual(ws['A17'].value, None, msg="not None when it should be")

        rem_file.unlink()


    def test_which_full_ramdom(self):
        test = TriangleTest()
        result, inputs = test.make_combinations(37)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(sheets[0], "sheet_1", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws['C6'].value, int(result[result.columns[1]].iloc[0][-3:]), msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, int(result[result.columns[2]].iloc[0][-3:]), msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, int(result[result.columns[3]].iloc[0][-3:]), msg="wrong 3digit value")

        rem_file.unlink()



    def test_which_full(self):
        test = TriangleTest()
        test.make_combinations(37)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)

        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 19, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "sheet_1", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws.max_row, 28, msg="row size does not match")
        self.assertEqual(ws['A1'].value, "Triangle Test", msg="wrong title")
        self.assertIsInstance(ws['C6'].value, int, msg="tri-digit not a number")
        self.assertIsInstance(ws['E6'].value, int, msg="tri-digit not a number")
        self.assertIsInstance(ws['G6'].value, int, msg="tri-digit not a number")
        rem_file.unlink()



    def test_which_full_output(self):
        test = TriangleTest()
        test.make_combinations(37)

        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput

        test.make_scoresheets(which=self.which_value)

        sys.stdout = sys.__stdout__
        expected = "he 'score_sheet.xlsx' file wa"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")
        rem_file.unlink()



class Test_make_scoresheets_which_full(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.which_value = "full"



    def test_which_full_seed_42(self):
        test = TriangleTest()
        test.make_combinations(37, seed=42)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)


        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "Triangle Test", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws['C6'].value, 817, msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, 493, msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, 480, msg="wrong 3digit value")

        self.assertEqual(ws['C51'].value, 212, msg="wrong 3digit value")
        self.assertEqual(ws['E51'].value, 168, msg="wrong 3digit value")
        self.assertEqual(ws['G51'].value, 296, msg="wrong 3digit value")


        self.assertEqual(ws['C111'].value, 668, msg="wrong 3digit value")
        self.assertEqual(ws['E111'].value, 355, msg="wrong 3digit value")
        self.assertEqual(ws['G111'].value, 918, msg="wrong 3digit value")

        self.assertEqual(ws['C291'].value, 910, msg="wrong 3digit value")
        self.assertEqual(ws['E291'].value, 665, msg="wrong 3digit value")
        self.assertEqual(ws['G291'].value, 312, msg="wrong 3digit value")

        result = False
        if "Panelist 20" in ws['A287'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")

        result = False
        if "Panelist 27" in ws['A392'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")

        result = False
        if "Panelist 7" in ws['A92'].value:
            result = True
        self.assertTrue(result, msg="wrong panelist")

        rem_file.unlink()



    def test_which_default_ramdom(self):
        test = TriangleTest()
        result, inputs = test.make_combinations(37)
        test.make_scoresheets()
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(sheets[0], "Triangle Test", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws['C6'].value, int(result[result.columns[1]].iloc[0][-3:]), msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, int(result[result.columns[2]].iloc[0][-3:]), msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, int(result[result.columns[3]].iloc[0][-3:]), msg="wrong 3digit value")

        rem_file.unlink()


    def test_which_full_ramdom(self):
        test = TriangleTest()
        result, inputs = test.make_combinations(37)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)
        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(sheets[0], "Triangle Test", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws['C6'].value, int(result[result.columns[1]].iloc[0][-3:]), msg="wrong 3digit value")
        self.assertEqual(ws['E6'].value, int(result[result.columns[2]].iloc[0][-3:]), msg="wrong 3digit value")
        self.assertEqual(ws['G6'].value, int(result[result.columns[3]].iloc[0][-3:]), msg="wrong 3digit value")

        rem_file.unlink()


    def test_which_full(self):
        test = TriangleTest()
        test.make_combinations(37)
        test.make_scoresheets(which=self.which_value)
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)


        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "Triangle Test", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws.max_row, 553, msg="row size does not match")
        self.assertEqual(ws['A1'].value, "Triangle Test", msg="wrong title")
        self.assertIsInstance(ws['C6'].value, int, msg="tri-digit not a number")
        self.assertIsInstance(ws['E6'].value, int, msg="tri-digit not a number")
        self.assertIsInstance(ws['G6'].value, int, msg="tri-digit not a number")
        rem_file.unlink()



    def test_which_full_output(self):
        test = TriangleTest()
        test.make_combinations(37)

        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput

        test.make_scoresheets(which="full")

        sys.stdout = sys.__stdout__
        expected = "e 'score_sheet.xlsx' fil"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")
        rem_file.unlink()
        ################


    def test_which_default(self):
        test = TriangleTest()
        test.make_combinations(37)
        test.make_scoresheets()
        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)


        wb = load_workbook(file_name)
        sheets = wb.sheetnames
        self.assertEqual(len(sheets), 1, msg="wrong number of sheets")
        self.assertEqual(sheets[0], "Triangle Test", msg="wrong sheet name")
        ws = wb[sheets[0]]
        self.assertEqual(ws.max_row, 553, msg="row size does not match")
        self.assertEqual(ws['A1'].value, "Triangle Test", msg="wrong title")
        self.assertIsInstance(ws['C6'].value, int, msg="tri-digit not a number")
        self.assertIsInstance(ws['E6'].value, int, msg="tri-digit not a number")
        self.assertIsInstance(ws['G6'].value, int, msg="tri-digit not a number")
        rem_file.unlink()


    def test_which_default_output(self):
        test = TriangleTest()
        test.make_combinations(37)

        file_name = "score_sheet.xlsx"
        rem_file = Path(file_name)

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput

        test.make_scoresheets()

        sys.stdout = sys.__stdout__
        expected = "e 'score_sheet.xlsx' fil"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")
        rem_file.unlink()
        ################



class Test_make_scoresheets(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        # criando um arquivo xlsx generico
        file_name = "bawitdaba.xlsx"
        # file_path = folder_name + "/" + file_name
        df = pd.DataFrame(columns=["a"], data=["c"])
        df.to_excel(file_name, index=False, engine="openpyxl")

        # fazendo o teste
        result = _check_file_exists(file_name)

        # para deletar o arquivo generico
        cls.rem_file = Path(file_name)


    @classmethod
    def tearDownClass(cls):
        cls.rem_file.unlink()

    def test_when_file_exists(self):
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(file_name="bawitdaba")
        result = _check_file_exists("bawitdaba.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("bawitdaba_1.xlsx")
        rem_file.unlink()

        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(file_name="bawitdaba")
        result = _check_file_exists("bawitdaba.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("bawitdaba_1.xlsx")
        rem_file.unlink()


    def test_when_file_exists_output(self):
        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(file_name="bawitdaba")
        rem_file = Path("bawitdaba_1.xlsx")
        rem_file.unlink()
        sys.stdout = sys.__stdout__
        expected = "UserWarning"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        expected = "'bawitdaba_1.xlsx' file"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        #################


        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(file_name="bawitdaba")
        rem_file = Path("bawitdaba_1.xlsx")
        rem_file.unlink()
        sys.stdout = sys.__stdout__
        expected = "UserWarning: O arquivo"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        expected = "quivo 'bawitdaba_1.xlsx'  f"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        #################


    def test_instructions_text(self):
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(which="one", instructions_text=["BABYMETAL RULES", "Metal", "Never, nnever"])
        result = _check_file_exists("score_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("score_sheet.xlsx")

        wb = load_workbook("score_sheet.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A5'].value, "BABYMETAL RULES", msg="wrong instructions_text")
        self.assertEqual(ws['A6'].value, "Metal", msg="wrong instructions_text")
        self.assertEqual(ws['A7'].value, "Never, nnever", msg="wrong instructions_text")
        rem_file.unlink()



        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(which="one", )
        result = _check_file_exists("planilha_score.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("planilha_score.xlsx")

        wb = load_workbook("planilha_score.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A5'].value, 'Avalie as amostras da esquerda para a direita. Duas são iguais. Marque um "X" na caixa da amostra que difere das demais. Se nenhuma diferença for aparente, você deve adivinhar.', msg="wrong instructions_text")
        rem_file.unlink()



        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(which="one", )
        result = _check_file_exists("score_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("score_sheet.xlsx")

        wb = load_workbook("score_sheet.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A5'].value, 'Evaluate samples from left to rigth. Two are alike. Mark an "X" in the box from the sample which differs from the others. If no difference is apparent, you must guess.', msg="wrong title")
        rem_file.unlink()




    def test_title(self):
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(which="one", title="BABYMETAL RULES")
        result = _check_file_exists("score_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("score_sheet.xlsx")

        wb = load_workbook("score_sheet.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A1'].value, "BABYMETAL RULES", msg="wrong title")
        rem_file.unlink()



        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(which="one", )
        result = _check_file_exists("planilha_score.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("planilha_score.xlsx")

        wb = load_workbook("planilha_score.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A1'].value, "Teste Triangular", msg="wrong title")
        rem_file.unlink()



        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets(which="one", )
        result = _check_file_exists("score_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("score_sheet.xlsx")

        wb = load_workbook("score_sheet.xlsx")
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        self.assertEqual(ws['A1'].value, "Triangle Test", msg="wrong title")
        rem_file.unlink()



    def test_pass(self):
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets()
        result = _check_file_exists("score_sheet.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("score_sheet.xlsx")
        rem_file.unlink()

        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets()
        result = _check_file_exists("planilha_score.xlsx")
        self.assertTrue(result, msg="Not True when file with same name is exported")
        rem_file = Path("planilha_score.xlsx")
        rem_file.unlink()


    def test_pass_output(self):
        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        test = TriangleTest()
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets()
        sys.stdout = sys.__stdout__
        expected = "The 'score_sheet.xlsx' file was expo"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")
        rem_file = Path("score_sheet.xlsx")
        rem_file.unlink()


        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        test = TriangleTest(language="pt-br")
        result, inpupt = test.make_combinations(37)
        test.make_scoresheets()
        sys.stdout = sys.__stdout__
        expected = "rquivo 'planilha_score.xlsx'  foi expor"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")
        rem_file = Path("planilha_score.xlsx")
        rem_file.unlink()


    def test_which_raises(self):
        test = TriangleTest()
        test.make_combinations(45)
        with self.assertRaises(ValueError, msg="Does not raised error when which is wrong"):
            test.make_scoresheets(which="a")
        with self.assertRaises(ValueError, msg="Does not raised error when which is wrong"):
            test.make_scoresheets(which=1)
        with self.assertRaises(ValueError, msg="Does not raised error when which is wrong"):
            test.make_scoresheets(which=["one"])
        with self.assertRaises(ValueError, msg="Does not raised error when which is wrong"):
            test.make_scoresheets(which="ful")


    def test_which_raises_output(self):
        test = TriangleTest()
        test.make_combinations(45)
        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        try:
            test.make_scoresheets(which="a")
        except ValueError:
            pass
        sys.stdout = sys.__stdout__
        expected = "-->    'a'"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        ################

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        try:
            test.make_scoresheets(which=1)
        except ValueError:
            pass
        sys.stdout = sys.__stdout__
        expected = "but we got a parameter of type 'int'"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        ################

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        try:
            test.make_scoresheets(which=["a"])
        except ValueError:
            pass
        sys.stdout = sys.__stdout__
        expected = "ut we got a parameter of type 'list'"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        ################

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        try:
            test.make_scoresheets(which="ful")
        except ValueError:
            pass
        sys.stdout = sys.__stdout__
        expected = "-->    'ful'"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        ################



    def test_raises_combinations_not_generated_yet(self):
        with self.assertRaises(ValueError, msg="Does not raised error when the combinatios were not generated yet"):
            test = TriangleTest()
            test.make_scoresheets()


    def test_raises_combinations_not_generated_yet_output(self):

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        try:
            test = TriangleTest()
            test.make_scoresheets()
        except ValueError:
            pass
        sys.stdout = sys.__stdout__
        expected = "the protocol it is necessary to generate the"
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")

        ################

        capturedOutput = io.StringIO()
        sys.stdout = capturedOutput
        try:
            test = TriangleTest(language="pt-br")
            test.make_scoresheets()
        except ValueError:
            pass
        sys.stdout = sys.__stdout__
        expected = "ilize o método 'make_combinations' para obter as "
        result = False
        if expected in capturedOutput.getvalue():
            result = True
        self.assertTrue(result, msg="wrong output")






# never an honest word, but thats when I run the world https://youtu.be/uNaHzwkDOIk?t=176

if __name__ == "__main__":
    unittest.main()
